import openpyxl
import csv

for num in range(65,91):
    tempxl = openpyxl.load_workbook(f"spreadsheet-{chr(num)}.xlsx")

    loadtempxl = tempxl.active
    csvfile=open(f"csv\spreadsheet-{chr(num)} sheet.csv", 'w')
    writer = csv.writer(csvfile)
    for row in range (loadtempxl.max_row):
        lst =[]
        for col in loadtempxl.iter_cols(1,loadtempxl.max_column):
            lst.append(col[row].value)
        writer.writerow(lst)
print('end')