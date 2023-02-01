import openpyxl,sys
wb = openpyxl.Workbook()
sheet = wb.active
n=int(sys.argv[1])
for r in range (1,n+2):
    for c in range(1,n+2):
        if c ==1 and r ==1:
            sheet.cell(r=1,c=1).value=' '
for r1 in range(2,n+2):
    sheet.cell(r=1,c=r1).value=r1-1
for c1 in range(2,n+2):
    sheet.cell(r=c1,c=1).value=c1-1
wb.save('mt.xlsx')

