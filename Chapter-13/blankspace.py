import sys, openpyxl
from openpyxl.utils import get_column_letter

if len(sys.argv) > 1:
	a = int(sys.argv[1])
	b = int(sys.argv[2])
	fileName = sys.argv[3]

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
wb2 = openpyxl.Workbook()
sheet2 = wb2.active

for i in range(1, sheet.max_row + 1):
	for j in range(1, sheet.max_column + 1):
		if i < a:
			cell = get_column_letter(j) + str(i)
		else:
			cell = get_column_letter(j) + str(i+b)
		sheet2[cell] = sheet.cell(row=i, column=j).value

wb2.save(f"{fileName}.xlsx")