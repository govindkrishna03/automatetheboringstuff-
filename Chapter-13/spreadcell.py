import  openpyxl


wb = openpyxl.load_workbook('produceSales.xlsx')
sheet1=wb.active
row=sheet1.max_row
column=sheet1.max_column

wb2=openpyxl.Workbook()
sheet2=wb2.active

for i in range(1,row+1):
        for j in range(1,column+1):
                sheet2.cell(row=j,column=i).value=sheet1.cell(row=i,column=j)
wb2.save('new_sheet.xlsx')
    
