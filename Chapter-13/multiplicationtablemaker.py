import openpyxl
import pyinputplus as pyip
from openpyxl.utils import get_column_letter


Number = pyip.inputNum('Please Enter a Number:')


wb = openpyxl.Workbook()
sheet = wb.active
for i in range(2, Number+2):
    sheet['A1'] = ''
    sheet['A'+str(i)] = i-1
    sheet[get_column_letter(i)+'1'] = i-1
    
    
wb.save('Multiplication.xlsx')