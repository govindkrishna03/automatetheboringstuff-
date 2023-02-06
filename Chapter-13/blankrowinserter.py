import sys, openpyxl
from openpyxl.utils import get_column_letter

if len(sys.argv) > 1:
	n = int(sys.argv[1])
	m = int(sys.argv[2])
	fileName = sys.argv[3]

wb = openpyxl.load_workbook(fileName)
sheet = wb.active
wb2 = openpyxl.Workbook()
sheet2 = wb2.active

for i in range(1, sheet.max_row + 1):
	for j in range(1, sheet.max_column + 1):
		if i < n:
			cell = get_column_letter(j) + str(i)
		else:
			cell = get_column_letter(j) + str(i+m)
		sheet2[cell] = sheet.cell(row=i, column=j).value

[baseName, ext] = fileName.split('.')
wb2.save(baseName + 'WithBlankRow.' + ext)